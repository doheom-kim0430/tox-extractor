import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import io
import os
import re

# ─────────────────────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="독성정보 자동 추출 시스템", layout="wide")
st.title("🧪 화학물질 독성정보 자동 추출 서비스")
st.info("내부식별자를 입력하면 DB에서 독성정보를 추출하여 엑셀 파일을 생성합니다.")

DB_FILENAME = "유해성미확인물질 12종 DB.xlsx"
TPL_SINGLE  = "개별물질 추출 템플릿.xlsx"
TPL_MULTI   = "다중물질 추출 템플릿.xlsx"

# ─────────────────────────────────────────────────────────────
# 개별물질 템플릿 열 매핑 (템플릿 직접 분석 기반)
# D=4  E=5  F=6  G=7  H=8  I=9  J=10  K=11  L=12  M=13  N=14  O=15  P=16  Q=17
# ECHA US   Pub  Kre  환경  TB_RA TB_Q  Dan   VEGA  Epi   HAZ   Pro   Vega  Chemi
# ─────────────────────────────────────────────────────────────
SINGLE_COLS = {
    'ECHA CHEM':            4,   # D
    'US DashBoard':         5,   # E
    'Pubchem':              6,   # F
    'K-reach':              7,   # G
    '환경부유해성심사결과': 8,   # H
    'TB_RA':                9,   # I  QSAR Toolbox Read-across
    'TB_QSAR':             10,   # J  QSAR Toolbox QSAR
    'Danish QSAR':         11,   # K
    'VEGA_QSAR':           12,   # L  VEGA QSAR
    'Epi suite':           13,   # M
    'HAZMAP':              14,   # N
    'Protox 3.0':          15,   # O
    'VEGA_AI':             16,   # P  VEGA AI-based QSAR
    'Cheminfomatics':      17,   # Q
}

# 개별물질 유해성 데이터 행 (row 12~21)
SINGLE_CAT_ROWS = {
    '급성경구독성':                        12,
    '급성흡입독성':                        13,
    '피부부식성/자극성':                   14,
    '복귀돌연변이':                        15,
    '포유류 배양세포를 이용한 염색체이상': 16,
    '소핵시험':                            17,
    '어류급성독성':                        18,
    '물벼룩급성독성':                      19,
    '담수조류생장저해':                    20,
    '이분해성':                            21,
}

# ─────────────────────────────────────────────────────────────
# 다중물질 템플릿 열 매핑 (템플릿 직접 분석 기반)
# F=6  G=7  H=8  I=9  J=10  K=11  L=12  M=13  N=14  O=15  P=16  Q=17  R=18  S=19
# ECHA US   Pub  Kre  환경  TB_RA  TB_Q  Dan   VEGA  Epi   HAZ   Pro   Vega  Chemi
# ─────────────────────────────────────────────────────────────
MULTI_COLS = {
    'ECHA CHEM':            6,   # F
    'US DashBoard':         7,   # G
    'Pubchem':              8,   # H
    'K-reach':              9,   # I
    '환경부유해성심사결과': 10,   # J
    'TB_RA':               11,   # K  QSAR Toolbox Read-across
    'TB_QSAR':             12,   # L  QSAR Toolbox QSAR
    'Danish QSAR':         13,   # M
    'VEGA_QSAR':           14,   # N  VEGA QSAR
    'Epi suite':           15,   # O
    'HAZMAP':              16,   # P
    'Protox 3.0':          17,   # Q
    'VEGA_AI':             18,   # R  VEGA AI-based QSAR
    'Cheminfomatics':      19,   # S
}

# 다중물질 블록 헤더행 (row2=블록1, row15=블록2)
MULTI_BLOCK_HEADERS = [2, 15]

# 유해성 항목 row offset (헤더행 기준)
MULTI_CAT_OFFSETS = {
    '급성경구독성':                         2,   # 블록1:row4,  블록2:row17
    '급성흡입독성':                         3,   # 블록1:row5,  블록2:row18
    '피부부식성/자극성':                    4,   # 블록1:row6,  블록2:row19
    '복귀돌연변이':                         5,   # 블록1:row7,  블록2:row20
    '포유류 배양세포를 이용한 염색체이상':  6,   # 블록1:row8,  블록2:row21
    '소핵시험':                             7,   # 블록1:row9,  블록2:row22
    '어류급성독성':                         8,   # 블록1:row10, 블록2:row23
    '물벼룩급성독성':                       9,   # 블록1:row11, 블록2:row24
    '담수조류생장저해':                    10,   # 블록1:row12, 블록2:row25
    '이분해성':                            11,   # 블록1:row13, 블록2:row26
}

# 물질정보 offset (헤더행 기준, INFO_COL=B=2)
MULTI_INFO_OFFSETS = {
    '내부식별자': 1,   # 블록1:row3,  블록2:row16
    'CAS No.':    3,   # 블록1:row5,  블록2:row18
    '물질명':     5,   # 블록1:row7,  블록2:row20
    '분자식':     7,   # 블록1:row9,  블록2:row22
    '분자량':     9,   # 블록1:row11, 블록2:row24
}
MULTI_INFO_COL = 2  # B열


# ─────────────────────────────────────────────────────────────
# 공통 유틸
# ─────────────────────────────────────────────────────────────

def write_safe(ws, row, col, value):
    """병합 셀 포함 안전 입력"""
    cell = ws.cell(row=row, column=col)
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            ws.cell(row=merged.min_row, column=merged.min_col).value = value
            return
    cell.value = value


# ─────────────────────────────────────────────────────────────
# VEGA 우선순위 로직
# ─────────────────────────────────────────────────────────────
VEGA_PRIORITY = ["EXPERIMENTAL value", "GOOD reliability", "MODERATE reliability", "LOW reliability"]

def get_best_vega(df):
    if df.empty:
        return None
    temp = df.copy()
    def rank(v):
        v = str(v)
        for i, label in enumerate(VEGA_PRIORITY):
            if label.lower() in v.lower():
                return len(VEGA_PRIORITY) - i
        return 0
    def score(v):
        m = re.search(r'\(([0-9.]+)\)', str(v))
        return float(m.group(1)) if m else 0.0
    temp['_rank']  = temp['Domain status'].apply(rank)
    temp['_score'] = temp['Domain status'].apply(score)
    return temp.sort_values(['_rank','_score'], ascending=[False,False]).iloc[0]


# ─────────────────────────────────────────────────────────────
# 공통 포맷 함수
# ─────────────────────────────────────────────────────────────

def apply_priority_exp(df, cat):
    """실험값 우선순위 정렬"""
    if len(df) <= 1:
        return df.iloc[0]
    temp = df.copy()
    if cat == "급성경구독성":
        temp['p1'] = (temp['Endpoint(표준)'] == 'LD50').astype(int)
        temp['p2'] = (temp['시험종(표준)'] == 'Rat').astype(int)
        temp['p3'] = temp['시험지침'].astype(str).str.contains('401', na=False).astype(int)
        temp = temp.sort_values(['p1','p2','p3','Result'], ascending=[False,False,False,True])
    elif cat == "급성흡입독성":
        temp['p1'] = (temp['Endpoint(표준)'] == 'LC50').astype(int)
        temp['p2'] = (temp['시험종(표준)'] == 'Rat').astype(int)
        temp['p3'] = (temp['Duration(표준)'] == '4 h').astype(int)
        temp['p4'] = temp['시험지침'].astype(str).str.contains('403', na=False).astype(int)
        temp = temp.sort_values(['p1','p2','p3','p4','Result'], ascending=[False,False,False,False,True])
    elif cat == "어류급성독성":
        temp['p1'] = (temp['Endpoint(표준)'] == 'LC50').astype(int)
        temp['p2'] = temp['시험종(표준)'].isin(['Fathead minnow','Zebrafish','Rainbow trout']).astype(int)
        temp['p3'] = (temp['Duration(표준)'] == '96 h').astype(int)
        temp['p4'] = temp['시험지침'].astype(str).str.contains('203', na=False).astype(int)
        temp = temp.sort_values(['p1','p2','p3','p4','Result'], ascending=[False,False,False,False,True])
    elif cat == "물벼룩급성독성":
        temp['p1'] = (temp['Endpoint(표준)'] == 'EC50').astype(int)
        temp['p2'] = (temp['시험종(표준)'] == 'Daphnia magna').astype(int)
        temp['p3'] = (temp['Duration(표준)'] == '48 h').astype(int)
        temp['p4'] = temp['시험지침'].astype(str).str.contains('202', na=False).astype(int)
        temp = temp.sort_values(['p1','p2','p3','p4','Result'], ascending=[False,False,False,False,True])
    elif cat == "담수조류생장저해":
        temp['p1'] = (temp['Endpoint(표준)'] == 'EC50').astype(int)
        temp['p2'] = temp['시험종(표준)'].isin(['P. subcapitata','D. subspicatus']).astype(int)
        temp['p3'] = (temp['Duration(표준)'] == '72 h').astype(int)
        temp['p4'] = temp['시험지침'].astype(str).str.contains('201', na=False).astype(int)
        temp = temp.sort_values(['p1','p2','p3','p4','Result'], ascending=[False,False,False,False,True])
    return temp.iloc[0]


def apply_priority_qsar_danish(df, cat, exp_species=None):
    """Danish QSAR 우선순위 정렬"""
    if len(df) <= 1:
        return df.iloc[0]
    temp = df.copy()
    model_map = {
        "급성경구독성":      "Acute toxicity in Rat, Oral - Danish QSAR DB ACDLabs model (v1.0)",
        "담수조류생장저해":  "Pseudokirchneriella subcapitata 72h EC50 - Danish QSAR DB battery model (v1.0)",
        "물벼룩급성독성":    "Daphnia magna 48h EC50 - Danish QSAR DB battery model (v1.0)",
        "복귀돌연변이":      "Ames test in S. typhimurium (in vitro) - Danish QSAR DB battery model (v1.0)",
        "소핵시험":          "Micronucleus Test in Mouse Erythrocytes - Danish QSAR DB battery model (v1.0)",
        "어류급성독성":      "Fathead minnow 96h LC50 - Danish QSAR DB battery model (v1.0)",
        "피부부식성/자극성": "BfR skin irritation/corrosion (v1.0)"
    }
    if cat == "포유류 배양세포를 이용한 염색체이상":
        mname = ("Chromosome Aberrations in Chinese Hamster Ovary (CHO) Cells - Danish QSAR DB battery model (v1.0)"
                 if exp_species == "CHO Cells"
                 else "Chromosome Aberrations in Chinese Hamster Lung (CHL) Cells - Danish QSAR DB battery model (v1.0)")
        temp['p_q'] = (temp['모델 종류 및 버전'] == mname).astype(int)
    elif cat in model_map:
        temp['p_q'] = (temp['모델 종류 및 버전'] == model_map[cat]).astype(int)
    else:
        temp['p_q'] = 0
    return temp.sort_values('p_q', ascending=False).iloc[0]


def format_exp(row, cat):
    """실험값/Read-across 포맷"""
    res = str(row['Result'])
    val_cats = ["급성경구독성","급성흡입독성","어류급성독성","물벼룩급성독성","담수조류생장저해"]
    if cat in val_cats:
        return f"{row['Endpoint(표준)']} = {res} {row['단위']} ({row['시험종(표준)']})"
    return res


def format_qsar(row, cat):
    """QSAR 포맷 (Out of domain 처리 포함)"""
    res = str(row['Result'])
    if str(row.get('Domain status','')) == "Out of domain":
        res += " (Out of domain)"
    val_cats = ["급성경구독성","급성흡입독성","어류급성독성","물벼룩급성독성","담수조류생장저해"]
    if cat in val_cats:
        return f"{row['Endpoint(표준)']} = {res} {row['단위']} ({row['시험종(표준)']})"
    return res


def format_multi_standard(row, cat):
    """다중 추출용 포맷"""
    res  = str(row['Result'])
    ep   = row.get('Endpoint') if pd.notna(row.get('Endpoint')) else (row.get('Endpoint(표준)','Unknown') or 'Unknown')
    sp   = row.get('시험종(표준)') if pd.notna(row.get('시험종(표준)')) else (row.get('시험종','Unknown') or 'Unknown')
    unit = row.get('단위','') if pd.notna(row.get('단위')) else ""
    if "(Out of domain)" not in res and pd.notna(row.get('Domain status')) and str(row.get('Domain status')) == "Out of domain":
        res += " (Out of domain)"
    val_cats = ["급성경구독성","급성흡입독성","어류급성독성","물벼룩급성독성","담수조류생장저해"]
    if cat in val_cats:
        return f"{ep} = {res} {unit} ({sp})"
    return res


def format_biodeg(row):
    """이분해성 포맷"""
    if row['출처'] in ['환경부유해성심사결과','K-reach'] or \
       (row['결과도출방법'] == 'QSAR' and row['출처'] == 'Epi suite'):
        return str(row['Result'])
    try:
        val = float(row['Result'])
        ep  = str(row['Endpoint']).lower()
        threshold = 70 if 'doc' in ep else 60
        status = "positive(이분해성)" if val >= threshold else "negative(난분해성)"
        return f"{status} - {row['Endpoint']} = {row['Result']} {row['단위']}"
    except:
        return str(row['Result'])


def filter_skin_exp(df):
    """피부부식성 실험값: Positive/Negative 중 Rabbit 우선"""
    temp = df[df['Result'].astype(str).str.lower().isin(['positive','negative'])]
    if not temp.empty:
        rabbit = temp[temp['시험종(표준)'].astype(str).str.contains('Rabbit', case=False, na=False)]
        return rabbit.iloc[0] if not rabbit.empty else temp.iloc[0]
    return None


def get_best_multi(df, cat):
    """다중 추출용 우선순위 (VEGA 제외 일반)"""
    if df.empty:
        return None
    temp = df.copy()
    temp['result_num'] = pd.to_numeric(temp['Result'], errors='coerce').fillna(999999)
    if cat == '이분해성':
        def gl(v):
            v = str(v).upper()
            return 2 if 'OECD' in v else (1 if v not in ['-','','NAN'] else 0)
        temp['gl'] = temp['시험지침'].apply(gl)
        return temp.sort_values(['gl','result_num'], ascending=[False,False]).iloc[0]
    if cat in ["급성경구독성","급성흡입독성","어류급성독성","물벼룩급성독성","담수조류생장저해"]:
        tep = "LD50" if "경구" in cat else ("LC50" if "어류" in cat or "흡입" in cat else "EC50")
        temp['ep_s'] = (
            temp['Endpoint'].astype(str).str.contains(tep, case=False, na=False) |
            temp['Endpoint(표준)'].astype(str).str.contains(tep, case=False, na=False)
        ).astype(int) * 10
        tsp = ("Rat" if "경구" in cat or "흡입" in cat else
               "Fathead minnow" if "어류" in cat else
               "Daphnia magna"  if "물벼룩" in cat else "P. subcapitata")
        temp['sp_s'] = temp['시험종(표준)'].astype(str).str.contains(tsp, case=False, na=False).astype(int) * 5
        temp['tot'] = temp['ep_s'] + temp['sp_s']
        return temp.sort_values(['tot','result_num'], ascending=[False,True]).iloc[0]
    return temp.iloc[0]


# ─────────────────────────────────────────────────────────────
# 단일 추출 실행
# ─────────────────────────────────────────────────────────────

def extract_single(target_id, df_mat, df_tox, wb):
    ws = wb.active

    # 물질 기본정보 (row7: C=내부식별자, D=CAS, E=물질명, F=분자식, G=분자량)
    mat_row = df_mat[df_mat['내부식별자'] == target_id]
    if mat_row.empty:
        raise ValueError(f"'{target_id}' 물질정보를 DB에서 찾을 수 없습니다.")
    t = mat_row.iloc[0]
    write_safe(ws, 7, 3, target_id)
    write_safe(ws, 7, 4, str(t['CAS']))
    write_safe(ws, 7, 5, str(t['물질명']))
    write_safe(ws, 7, 6, str(t['분자식']))
    write_safe(ws, 7, 7, str(t['분자량']))

    for cat, data_row in SINGLE_CAT_ROWS.items():
        df_cat = df_tox[(df_tox['내부식별자'] == target_id) & (df_tox['유해성항목'] == cat)]
        exp_species_found = None

        # ── 실험값 (D~H, col 4~8) ──
        for src, col in [('ECHA CHEM',4),('US DashBoard',5),('Pubchem',6),('K-reach',7),('환경부유해성심사결과',8)]:
            df_s = df_cat[(df_cat['결과도출방법'] == '실험값') & (df_cat['출처'] == src)]
            if df_s.empty:
                continue
            if cat == '피부부식성/자극성':
                best = filter_skin_exp(df_s)
            else:
                best = apply_priority_exp(df_s, cat)
            if best is not None:
                ws.cell(row=data_row, column=col).value = format_exp(best, cat)
                if cat == "포유류 배양세포를 이용한 염색체이상":
                    exp_species_found = best['시험종(표준)']

        # ── QSAR Toolbox Read-across (I=9) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'Read-across') & (df_cat['출처'] == 'QSAR Toolbox v.4.8')]
        if not df_s.empty:
            ws.cell(row=data_row, column=9).value = format_multi_standard(df_s.iloc[0], cat)

        # ── QSAR Toolbox QSAR (J=10) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'QSAR') & (df_cat['출처'] == 'QSAR Toolbox v.4.8')]
        if not df_s.empty:
            ws.cell(row=data_row, column=10).value = format_qsar(df_s.iloc[0], cat)

        # ── Danish QSAR (K=11) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'QSAR') & (df_cat['출처'] == 'Danish QSAR')]
        if not df_s.empty:
            best = apply_priority_qsar_danish(df_s, cat, exp_species_found)
            ws.cell(row=data_row, column=11).value = format_qsar(best, cat)

        # ── VEGA QSAR (L=12) ──
        df_s = df_cat[(df_cat['출처'] == 'VEGA') & (df_cat['결과도출방법'] == 'QSAR')]
        if not df_s.empty:
            best = get_best_vega(df_s)
            if best is not None:
                ws.cell(row=data_row, column=12).value = format_qsar(best, cat)

        # ── Epi suite (M=13) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'QSAR') & (df_cat['출처'] == 'Epi suite')]
        if not df_s.empty:
            if cat == '이분해성':
                ws.cell(row=data_row, column=13).value = format_biodeg(df_s.iloc[0])
            else:
                ws.cell(row=data_row, column=13).value = format_qsar(df_s.iloc[0], cat)

        # ── HAZMAP (N=14) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'AI-based QSAR') & (df_cat['출처'] == 'HAZMAP')]
        if not df_s.empty:
            ws.cell(row=data_row, column=14).value = str(df_s.iloc[0]['Result'])

        # ── Protox 3.0 (O=15) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'AI-based QSAR') & (df_cat['출처'] == 'Protox 3.0')]
        if not df_s.empty:
            ws.cell(row=data_row, column=15).value = str(df_s.iloc[0]['Result'])

        # ── VEGA AI-based QSAR (P=16) ──
        df_s = df_cat[(df_cat['출처'] == 'VEGA') & (df_cat['결과도출방법'] == 'AI-based QSAR')]
        if not df_s.empty:
            best = get_best_vega(df_s)
            if best is not None:
                ws.cell(row=data_row, column=16).value = str(best['Result'])

        # ── Cheminfomatics (Q=17) ──
        df_s = df_cat[(df_cat['결과도출방법'] == 'AI-based QSAR') & (df_cat['출처'] == 'Cheminfomatics')]
        if not df_s.empty:
            ws.cell(row=data_row, column=17).value = str(df_s.iloc[0]['Result'])

    # 스타일
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
    for rng in [ws['C7:G7'], ws['B11:Q21']]:
        for row in rng:
            for cell in row:
                cell.border    = thin
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font      = Font(name='맑은 고딕', size=9)
    col_widths = {'B':12,'C':15,'D':22,'E':25,'F':12,'G':12,'H':22,
                  'I':18,'J':20,'K':20,'L':20,'M':20,'N':15,'O':15,'P':15,'Q':15}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for i in range(12, 22):
        ws.row_dimensions[i].height = 45


# ─────────────────────────────────────────────────────────────
# 다중 추출 실행
# ─────────────────────────────────────────────────────────────

def extract_multi(tid1, tid2, df_mat, df_tox, wb):
    ws = wb.active
    ws.title = f"{tid1} 및 {tid2}"

    # 데이터 셀 초기화
    for hdr in MULTI_BLOCK_HEADERS:
        for offset in MULTI_INFO_OFFSETS.values():
            ws.cell(row=hdr + offset, column=MULTI_INFO_COL).value = None
        for offset in MULTI_CAT_OFFSETS.values():
            for col in range(6, 20):
                ws.cell(row=hdr + offset, column=col).value = None

    for tid, hdr_row in zip([tid1, tid2], MULTI_BLOCK_HEADERS):
        mat_row = df_mat[df_mat['내부식별자'] == tid]
        if mat_row.empty:
            raise ValueError(f"'{tid}' 물질정보를 DB에서 찾을 수 없습니다.")
        t = mat_row.iloc[0]

        # 물질 기본정보
        for label, offset in MULTI_INFO_OFFSETS.items():
            val = {
                '내부식별자': tid,
                'CAS No.':    str(t['CAS']),
                '물질명':     str(t['물질명']),
                '분자식':     str(t['분자식']),
                '분자량':     f"{t['분자량']} g/mol",
            }[label]
            write_safe(ws, hdr_row + offset, MULTI_INFO_COL, val)

        df_sub = df_tox[df_tox['내부식별자'] == tid]

        for cat, cat_offset in MULTI_CAT_OFFSETS.items():
            data_row = hdr_row + cat_offset
            df_cat   = df_sub[df_sub['유해성항목'] == cat]

            # ── 실험값 (F~J, col 6~10) ──
            for src, col in [('ECHA CHEM',6),('US DashBoard',7),('Pubchem',8),('K-reach',9),('환경부유해성심사결과',10)]:
                df_s = df_cat[(df_cat['결과도출방법'] == '실험값') & (df_cat['출처'] == src)]
                if df_s.empty:
                    continue
                if cat == '피부부식성/자극성':
                    best = filter_skin_exp(df_s)
                else:
                    best = apply_priority_exp(df_s, cat)
                if best is not None:
                    write_safe(ws, data_row, col, format_multi_standard(best, cat))

            # ── QSAR Toolbox Read-across (K=11) ──
            df_s = df_cat[
                df_cat['출처'].astype(str).str.contains('QSAR Toolbox', case=False, na=False) &
                df_cat['결과도출방법'].astype(str).str.contains('Read across', case=False, na=False)
            ]
            if not df_s.empty:
                write_safe(ws, data_row, 11, format_multi_standard(df_s.iloc[0], cat))

            # ── QSAR Toolbox QSAR (L=12) ──
            df_s = df_cat[
                df_cat['출처'].astype(str).str.contains('QSAR Toolbox', case=False, na=False) &
                ~df_cat['결과도출방법'].astype(str).str.contains('Read across', case=False, na=False) &
                (df_cat['결과도출방법'] == 'QSAR')
            ]
            if not df_s.empty:
                write_safe(ws, data_row, 12, format_multi_standard(df_s.iloc[0], cat))

            # ── Danish QSAR (M=13) ──
            df_s = df_cat[(df_cat['결과도출방법'] == 'QSAR') & (df_cat['출처'] == 'Danish QSAR')]
            if not df_s.empty:
                best = get_best_multi(df_s, cat)
                if best is not None:
                    write_safe(ws, data_row, 13, format_multi_standard(best, cat))

            # ── VEGA QSAR (N=14) ──
            df_s = df_cat[(df_cat['출처'] == 'VEGA') & (df_cat['결과도출방법'] == 'QSAR')]
            if not df_s.empty:
                best = get_best_vega(df_s)
                if best is not None:
                    write_safe(ws, data_row, 14, format_multi_standard(best, cat))

            # ── Epi suite (O=15) ──
            df_s = df_cat[(df_cat['결과도출방법'] == 'QSAR') & (df_cat['출처'] == 'Epi suite')]
            if not df_s.empty:
                if cat == '이분해성':
                    write_safe(ws, data_row, 15, format_biodeg(df_s.iloc[0]))
                else:
                    write_safe(ws, data_row, 15, format_multi_standard(df_s.iloc[0], cat))

            # ── HAZMAP (P=16) ──
            df_s = df_cat[(df_cat['결과도출방법'] == 'AI-based QSAR') & (df_cat['출처'] == 'HAZMAP')]
            if not df_s.empty:
                write_safe(ws, data_row, 16, str(df_s.iloc[0]['Result']))

            # ── Protox 3.0 (Q=17) ──
            df_s = df_cat[(df_cat['결과도출방법'] == 'AI-based QSAR') & (df_cat['출처'] == 'Protox 3.0')]
            if not df_s.empty:
                write_safe(ws, data_row, 17, str(df_s.iloc[0]['Result']))

            # ── VEGA AI-based QSAR (R=18) ──
            df_s = df_cat[(df_cat['출처'] == 'VEGA') & (df_cat['결과도출방법'] == 'AI-based QSAR')]
            if not df_s.empty:
                best = get_best_vega(df_s)
                if best is not None:
                    write_safe(ws, data_row, 18, str(best['Result']))

            # ── Cheminfomatics (S=19) ──
            df_s = df_cat[(df_cat['결과도출방법'] == 'AI-based QSAR') & (df_cat['출처'] == 'Cheminfomatics')]
            if not df_s.empty:
                write_safe(ws, data_row, 19, str(df_s.iloc[0]['Result']))

    # 스타일
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
    for hdr in MULTI_BLOCK_HEADERS:
        for r in range(hdr + 2, hdr + 12):
            for c in range(6, 20):
                cell = ws.cell(row=r, column=c)
                cell.border    = thin
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font      = Font(name='맑은 고딕', size=9)


# ─────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────

if not os.path.exists(DB_FILENAME):
    st.error(f"DB 파일을 찾을 수 없습니다: **{DB_FILENAME}**")
    st.stop()

mode = st.radio("📋 추출 모드 선택", ["단일 물질 추출", "다중 물질 추출 (2개)"], horizontal=True)
st.divider()

if mode == "단일 물질 추출":
    if not os.path.exists(TPL_SINGLE):
        st.error(f"템플릿 파일 없음: **{TPL_SINGLE}**")
        st.stop()
    target_id = st.text_input("🔍 내부식별자 입력 (예: B-3)", value="B-3")
    if st.button("🚀 추출 및 엑셀 다운로드", key="btn_single"):
        with st.spinner("데이터 추출 중..."):
            try:
                df_mat = pd.read_excel(DB_FILENAME, sheet_name='물질정보')
                df_tox = pd.read_excel(DB_FILENAME, sheet_name='유해성정보')
                wb     = load_workbook(TPL_SINGLE)
                extract_single(target_id.strip(), df_mat, df_tox, wb)
                buf = io.BytesIO()
                wb.save(buf)
                st.success(f"✅ **{target_id}** 추출 완료!")
                st.download_button(
                    label="📥 결과 엑셀 다운로드",
                    data=buf.getvalue(),
                    file_name=f"추출결과_{target_id}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"오류 발생: {e}")

else:
    if not os.path.exists(TPL_MULTI):
        st.error(f"템플릿 파일 없음: **{TPL_MULTI}**")
        st.stop()
    col1, col2 = st.columns(2)
    with col1:
        tid1 = st.text_input("🔍 첫 번째 내부식별자 (예: B-1)", value="B-1")
    with col2:
        tid2 = st.text_input("🔍 두 번째 내부식별자 (예: B-3)", value="B-3")
    if st.button("🚀 추출 및 엑셀 다운로드", key="btn_multi"):
        if not tid1.strip() or not tid2.strip():
            st.warning("두 개의 내부식별자를 모두 입력해주세요.")
        elif tid1.strip() == tid2.strip():
            st.warning("서로 다른 내부식별자를 입력해주세요.")
        else:
            with st.spinner("데이터 추출 중..."):
                try:
                    df_mat = pd.read_excel(DB_FILENAME, sheet_name='물질정보')
                    df_tox = pd.read_excel(DB_FILENAME, sheet_name='유해성정보')
                    wb     = load_workbook(TPL_MULTI)
                    extract_multi(tid1.strip(), tid2.strip(), df_mat, df_tox, wb)
                    buf = io.BytesIO()
                    wb.save(buf)
                    st.success(f"✅ **{tid1}** + **{tid2}** 추출 완료!")
                    st.download_button(
                        label="📥 결과 엑셀 다운로드",
                        data=buf.getvalue(),
                        file_name=f"추출결과_{tid1}_{tid2}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"오류 발생: {e}")
